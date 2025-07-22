import streamlit as st
import openpyxl
from openpyxl import Workbook
import re
from io import BytesIO
import pandas as pd

# ---------------- Streamlit Config ----------------
st.set_page_config(
    page_title="CPA Extractor Pro",
    page_icon="📈",
    layout="centered"
)

st.markdown("""
    <style>
    .main {
        background-color: #f4f6f9;
    }
    .stApp {
        font-family: 'Segoe UI', sans-serif;
    }
    .uploadbox {
        background-color: #ffffff;
        padding: 1.5em;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    .resultbox {
        background-color: #fefefe;
        padding: 1.5em;
        border-radius: 10px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.08);
        margin-top: 1em;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown("## 📄 CPA Excel Extractor Pro")

# ---------------- Helper Functions ----------------
def get_merged_cell_value(sheet, cell_coord):
    for merged_range in sheet.merged_cells.ranges:
        if cell_coord in merged_range:
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return sheet[cell_coord].value

def get_numeric_or_formula_value(sheet, cell_coord):
    value = get_merged_cell_value(sheet, cell_coord)
    if isinstance(value, (int, float)):
        return round(value)
    elif isinstance(value, str) and value.startswith('='):
        match = re.match(r"=SUM\((\w)(\d+):(\w)(\d+)\)", value, re.IGNORECASE)
        if match:
            col1, row1, col2, row2 = match.groups()
            total = 0.0
            for row in range(int(row1), int(row2) + 1):
                cell_val = get_merged_cell_value(sheet, f"{col1}{row}")
                if isinstance(cell_val, (int, float)):
                    total += cell_val
            return round(total)
    return "ERROR: Not numeric or unhandled formula"

def clean_customer_name(name):
    if not isinstance(name, str):
        return name
    cleaned = re.sub(r'\(?PVT\)?\.?\s*LTD', '', name, flags=re.IGNORECASE)
    return cleaned.strip()

def find_total_qty_row(sheet):
    for row in range(13, 100):
        b_val = get_merged_cell_value(sheet, f'B{row}')
        if isinstance(b_val, str) and 'total qty' in b_val.lower():
            return row
    return None

def get_last_numeric_in_column(sheet, col, reference_row):
    val = get_merged_cell_value(sheet, f'{col}{reference_row}')
    if isinstance(val, (int, float)):
        return round(val)
    elif isinstance(val, str) and val.startswith('='):
        return get_numeric_or_formula_value(sheet, f'{col}{reference_row}')
    else:
        return "ERROR: TOTAL QTY cell not numeric"

def extract_total_value(sheet):
    for row in range(24, sheet.max_row + 1):
        b_val = get_merged_cell_value(sheet, f'B{row}')
        if isinstance(b_val, str) and "total value usd" in b_val.lower():
            return get_numeric_or_formula_value(sheet, f'J{row}')
    return 'ERROR: Total value USD not found'

def extract_wo_numbers(sheet, end_row):
    wo_numbers = []
    for row in range(13, end_row):
        b_val = get_merged_cell_value(sheet, f'B{row}')
        if b_val and isinstance(b_val, str) and b_val.strip().startswith('SW'):
            wo_numbers.append(b_val.strip())
    return wo_numbers if wo_numbers else ['ERROR: No WO found']

def extract_references(sheet, start_row=13):
    references = []
    for row in range(start_row, sheet.max_row + 1):
        c_val = get_merged_cell_value(sheet, f'C{row}')
        if c_val is None or str(c_val).strip() == '':
            break
        references.append(str(c_val).strip())
    return references if references else ['ERROR: No Reference found']

def extract_reason_and_remarks(sheet):
    reason_row = None
    for row in range(15, sheet.max_row + 1):
        b_val = get_merged_cell_value(sheet, f'B{row}')
        if isinstance(b_val, str) and 'reason' in b_val.lower():
            reason_row = row
            break

    if not reason_row:
        return 'ERROR: Reason not found', 'ERROR: Remarks not found'

    reason_text = get_merged_cell_value(sheet, f'C{reason_row}') or get_merged_cell_value(sheet, f'B{reason_row}') or 'ERROR: Empty Reason'
    stop_keywords = ['artwork', 'printed', 'packed', 'quality', 'platemaking', 'created', 'cut']
    remarks = []

    for row in range(reason_row + 1, sheet.max_row + 1):
        b_val = get_merged_cell_value(sheet, f'B{row}')
        c_val = get_merged_cell_value(sheet, f'C{row}')
        if b_val is None and c_val is None:
            continue
        b_str = str(b_val).strip().lower() if b_val else ''
        c_str = str(c_val).strip() if c_val else ''
        if any(kw in b_str for kw in stop_keywords) and c_str:
            remarks.append(f"{b_val.strip() if b_val else 'Unknown'}: {c_str}")

    remarks_text = ', '.join(remarks) if remarks else 'ERROR: No remarks found'
    return reason_text, remarks_text

def extract_data_from_sheet(sheet):
    data = {}
    data['Date'] = get_merged_cell_value(sheet, 'D9') or 'ERROR: Missing D9'
    data['CPA No'] = get_merged_cell_value(sheet, 'H8') or 'ERROR: Missing H8'
    raw_customer = get_merged_cell_value(sheet, 'D10') or 'ERROR: Missing D10'
    data['Customer'] = clean_customer_name(raw_customer)
    total_qty_row = find_total_qty_row(sheet) or sheet.max_row
    data['WO'] = ', '.join(extract_wo_numbers(sheet, total_qty_row))
    data['Reference'] = ', '.join(extract_references(sheet))
    reason, remarks = extract_reason_and_remarks(sheet)
    data['Cause of Rejection'] = reason
    data['Remarks'] = remarks
    data['QTY'] = get_last_numeric_in_column(sheet, 'I', total_qty_row)
    data['MTRS'] = get_last_numeric_in_column(sheet, 'J', total_qty_row)
    data['Value'] = extract_total_value(sheet)
    return data

def process_workbook_from_stream(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "Extracted CPA Data"
    headers = ['Sheet Name', 'Date', 'CPA No', 'Customer', 'WO', 'Reference',
               'Cause of Rejection', 'QTY', 'MTRS', 'Value', 'Remarks']
    output_ws.append(headers)

    display_rows = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet.sheet_state != "visible":
            continue  # Skip hidden sheets

        extracted = extract_data_from_sheet(sheet)
        row = [sheet_name] + [extracted.get(h, '') for h in headers[1:]]
        output_ws.append(row)
        display_rows.append(row)

    output_stream = BytesIO()
    output_wb.save(output_stream)
    output_stream.seek(0)
    df_preview = pd.DataFrame(display_rows, columns=headers)
    return output_stream, df_preview


# ----------------- Upload & Display UI -----------------
with st.container():
    with st.expander("📤 Upload Excel File (.xlsx)", expanded=True):
        uploaded_file = st.file_uploader("Choose a file", type=["xlsx"], label_visibility="collapsed")

    if uploaded_file:
        with st.spinner("🔄 Processing... Please wait..."):
            excel_data, preview_df = process_workbook_from_stream(uploaded_file)
            st.success("✅ Data successfully extracted!")

            with st.container():
                st.markdown("### 🧾 Preview Extracted Data")
                st.dataframe(preview_df, use_container_width=True)

            st.download_button(
                label="📥 Download Processed Excel File",
                data=excel_data,
                file_name="Extracted_CPA.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("⬆ Upload your Excel file to begin.")